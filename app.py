import os
import io
import json
import time
import base64
import secrets
import threading
import hashlib
import hmac
from datetime import datetime, timezone
from functools import wraps
from urllib.parse import quote

import qrcode
import requests
from openpyxl import load_workbook
from cryptography.fernet import Fernet, InvalidToken
from flask import (
    Flask, request, redirect, url_for, session, render_template_string,
    jsonify, send_file, abort
)

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 12 * 1024 * 1024

ADMIN_PIN = os.environ.get("ADMIN_PIN", "")
DATA_KEY = os.environ.get("JCRC_MASTER_KEY", "")
GITHUB_TOKEN = os.environ.get("GITHUB_DATA_TOKEN", "")
app.secret_key = hashlib.sha256(DATA_KEY.encode("utf-8")).hexdigest() if DATA_KEY else "CHANGE-ME"
app.config["PERMANENT_SESSION_LIFETIME"] = 60 * 60 * 12
GITHUB_REPO = os.environ.get(
    "GITHUB_REPO", "secretariajockeyclub-glitch/vouchers-pileta-jcrc"
)
GITHUB_BRANCH = os.environ.get("GITHUB_BRANCH", "main")
STATE_PATH = os.environ.get("STATE_PATH", "state.json")
BASE_URL = os.environ.get("BASE_URL", "").rstrip("/")
VERIFY_TOKEN = os.environ.get("VERIFY_TOKEN", "jcrc_vouchers_2026")

_state_lock = threading.Lock()


def now_iso():
    return datetime.now(timezone.utc).isoformat()


def now_ts():
    return int(time.time())


def state_url():
    return f"https://api.github.com/repos/{GITHUB_REPO}/contents/{STATE_PATH}"


def gh_headers():
    return {
        "Authorization": f"Bearer {GITHUB_TOKEN}",
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
        "User-Agent": "Vouchers-Pileta-JCRC",
    }


def read_local_state():
    with open(STATE_PATH, "r", encoding="utf-8") as f:
        return json.load(f), None


def load_state():
    if not GITHUB_TOKEN:
        return read_local_state()
    r = requests.get(
        state_url(),
        headers=gh_headers(),
        params={"ref": GITHUB_BRANCH},
        timeout=15,
    )
    r.raise_for_status()
    obj = r.json()
    raw = base64.b64decode(obj["content"]).decode("utf-8")
    return json.loads(raw), obj["sha"]


def save_state(state, sha=None):
    data = json.dumps(state, ensure_ascii=False, indent=2).encode("utf-8")
    if not GITHUB_TOKEN:
        with open(STATE_PATH, "wb") as f:
            f.write(data)
        return

    if sha is None:
        _, sha = load_state()

    payload = {
        "message": "Actualiza estado de vouchers [skip render]",
        "content": base64.b64encode(data).decode("ascii"),
        "branch": GITHUB_BRANCH,
        "sha": sha,
    }
    r = requests.put(
        state_url(),
        headers=gh_headers(),
        json=payload,
        timeout=20,
    )
    if r.status_code not in (200, 201):
        raise RuntimeError(f"GitHub state save failed: {r.status_code} {r.text[:300]}")


def update_state(mutator):
    with _state_lock:
        last_error = None
        for _ in range(3):
            try:
                state, sha = load_state()
                result = mutator(state)
                save_state(state, sha)
                return result
            except Exception as e:
                last_error = e
                time.sleep(0.5)
        raise last_error


def fernet():
    if not DATA_KEY:
        raise RuntimeError("Falta JCRC_MASTER_KEY en Render.")
    return Fernet(DATA_KEY.encode("utf-8"))


def decrypt_member(member):
    try:
        raw = fernet().decrypt(member["payload"].encode("utf-8"))
        return json.loads(raw.decode("utf-8"))
    except (InvalidToken, KeyError) as e:
        raise RuntimeError("No se pudo leer la ficha. Revisá DATA_ENCRYPTION_KEY.") from e


def encrypt_payload(payload):
    raw = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    return fernet().encrypt(raw).decode("utf-8")


def normalize_phone(value):
    d = "".join(ch for ch in str(value or "") if ch.isdigit())
    if d.startswith("00"):
        d = d[2:]
    if d.startswith("54"):
        rest = d[2:].lstrip("0")
        if rest.startswith("9"):
            return "54" + rest
        return "549" + rest
    d = d.lstrip("0")
    if len(d) == 9 and d.startswith("15"):
        d = "358" + d[2:]
    if len(d) == 10:
        return "549" + d
    return ""



def clean_excel_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def clean_excel_int(value, default=0):
    if value is None or value == "":
        return default
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def clean_excel_date(value):
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    return clean_excel_text(value)


def member_match_key(payload):
    socio = clean_excel_text(payload.get("socio", "")).lower()
    if socio:
        return "socio:" + socio
    nombre = " ".join(clean_excel_text(payload.get("nombre", "")).lower().split())
    excel_id = clean_excel_text(payload.get("excel_id", ""))
    return f"fallback:{excel_id}:{nombre}"


def parse_excel_members(file_storage):
    """
    Lee A:H de forma secuencial. Esto evita accesos celda-por-celda en
    openpyxl read_only, que en plan Free de Render podían tardar más de
    40 segundos y provocar Internal Server Error.
    """
    raw = file_storage.read()
    if not raw:
        raise ValueError("El archivo está vacío.")
    try:
        wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    except Exception as e:
        raise ValueError("No se pudo abrir el Excel. Usá un archivo .xlsx válido.") from e

    ws = wb[wb.sheetnames[0]]

    # Leer A:H de una sola pasada es muchísimo más rápido que ws.cell(...)
    all_rows = list(ws.iter_rows(min_col=1, max_col=8, values_only=True))

    header_index = None
    for i, vals in enumerate(all_rows[:25]):
        c4 = clean_excel_text(vals[3] if len(vals) > 3 else "").lower()
        c7 = clean_excel_text(vals[6] if len(vals) > 6 else "").lower()
        if ("apellido" in c4 or "nombre" in c4) and "invit" in c7:
            header_index = i
            break

    if header_index is None:
        header_index = 3  # fila 4, como respaldo

    rows = []
    seen = set()

    for vals in all_rows[header_index + 1:]:
        # iter_rows ya devuelve exactamente 8 columnas, pero lo dejamos robusto
        vals = list(vals) + [None] * (8 - len(vals))

        nombre = clean_excel_text(vals[3])
        socio = clean_excel_text(vals[2])

        # Ignorar filas vacías o sólo formateadas
        if not nombre and not socio:
            continue

        invitaciones = max(0, clean_excel_int(vals[6], 0))
        payload = {
            "excel_id": clean_excel_text(vals[0]),
            "fecha": clean_excel_date(vals[1]),
            "socio": socio,
            "nombre": nombre,
            "categoria": clean_excel_text(vals[4]),
            "cantidad": max(0, clean_excel_int(vals[5], 0)),
            "telefono_original": clean_excel_text(vals[7]),
            "telefono_wa": normalize_phone(vals[7]),
        }

        key = member_match_key(payload)
        if key in seen:
            raise ValueError(
                f"Hay un Nº de socio duplicado en el Excel: {socio or nombre}"
            )
        seen.add(key)
        rows.append((key, payload, invitaciones))

    if not rows:
        raise ValueError("No encontré titulares en las columnas A:H del Excel.")

    return rows


def sync_excel_into_state(state, excel_rows):
    """
    El Excel es la fuente de datos maestros A:H.

    La columna Invitaciones (G) se usa como una CARGA PENDIENTE:
    - En un titular nuevo, el valor inicial se acredita al saldo.
    - Cuando luego el Excel pasa a 0, el saldo existente NO se borra.
    - Si más adelante pasa de 0 a un valor mayor que 0, ese valor se
      suma una sola vez al saldo como una nueva compra/carga.
    - Repetir una actualización con el mismo valor no duplica saldo.
    """
    current = state.setdefault("members", {})
    by_key = {}
    for mid, m in current.items():
        try:
            p = decrypt_member(m)
            by_key[member_match_key(p)] = (mid, m, p)
        except Exception:
            continue

    active_ids = set()
    added = updated = credited = 0
    credited_qty = 0

    for key, payload, excel_invitaciones in excel_rows:
        found = by_key.get(key)

        if found:
            mid, m, old_payload = found

            # Actualiza SIEMPRE todos los datos maestros del Excel.
            m["payload"] = encrypt_payload(payload)
            m["activo"] = True

            saldo = max(0, int(m.get("saldo", 0)))
            total_cargadas = max(
                0,
                int(m.get("invitaciones_iniciales", saldo))
            )

            # Compatibilidad / migración desde versiones anteriores.
            # Si todavía no existe excel_invitaciones_last:
            # - saldo > 0: tomamos el valor actual como ya conocido, sin duplicar.
            # - saldo == 0 y Excel > 0: lo interpretamos como una NUEVA COMPRA
            #   pendiente y la acreditamos una sola vez.
            if "excel_invitaciones_last" not in m:
                if saldo == 0 and excel_invitaciones > 0:
                    saldo += excel_invitaciones
                    total_cargadas += excel_invitaciones
                    credited += 1
                    credited_qty += excel_invitaciones
                    state.setdefault("history", []).append({
                        "at": now_iso(),
                        "type": "excel_credit_migration",
                        "member_id": mid,
                        "qty": excel_invitaciones,
                        "saldo": saldo,
                    })
                m["excel_invitaciones_last"] = excel_invitaciones
            else:
                last_excel = max(0, int(m.get("excel_invitaciones_last", 0)))

                # Una nueva carga se reconoce únicamente en la transición 0 -> N.
                if last_excel == 0 and excel_invitaciones > 0:
                    saldo += excel_invitaciones
                    total_cargadas += excel_invitaciones
                    credited += 1
                    credited_qty += excel_invitaciones
                    state.setdefault("history", []).append({
                        "at": now_iso(),
                        "type": "excel_credit",
                        "member_id": mid,
                        "qty": excel_invitaciones,
                        "saldo": saldo,
                    })

                # Guardamos el valor visto para evitar acreditarlo dos veces.
                m["excel_invitaciones_last"] = excel_invitaciones

            m["saldo"] = saldo
            m["invitaciones_iniciales"] = total_cargadas
            updated += 1

        else:
            mid = secrets.token_urlsafe(12)
            while mid in current:
                mid = secrets.token_urlsafe(12)

            current[mid] = {
                "payload": encrypt_payload(payload),
                "saldo": excel_invitaciones,
                "invitaciones_iniciales": excel_invitaciones,
                "excel_invitaciones_last": excel_invitaciones,
                "activo": True,
            }
            added += 1

        active_ids.add(mid)

    deactivated = 0
    for mid, m in current.items():
        if mid not in active_ids and m.get("activo", True):
            m["activo"] = False
            deactivated += 1

    state.setdefault("history", []).append({
        "at": now_iso(),
        "type": "excel_sync",
        "rows": len(excel_rows),
        "added": added,
        "updated": updated,
        "credited_members": credited,
        "credited_qty": credited_qty,
        "deactivated": deactivated,
    })
    state["history"] = state["history"][-500:]

    return {
        "rows": len(excel_rows),
        "added": added,
        "updated": updated,
        "credited_members": credited,
        "credited_qty": credited_qty,
        "deactivated": deactivated,
    }


def public_base():
    return BASE_URL or request.url_root.rstrip("/")


def voucher_share_signature(mid):
    """Firma corta y estable para compartir el voucher sin exponer el panel."""
    secret = DATA_KEY or app.secret_key
    return hmac.new(
        secret.encode("utf-8"),
        mid.encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()[:24]


def voucher_share_url(mid):
    return f"{public_base()}/voucher/{mid}/{voucher_share_signature(mid)}"


def valid_voucher_signature(mid, sig):
    return hmac.compare_digest(voucher_share_signature(mid), sig)


def is_admin():
    return bool(session.get("admin"))


def admin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not is_admin():
            session["next"] = request.url
            return redirect(url_for("login"))
        return fn(*args, **kwargs)
    return wrapper


BASE_CSS = """
:root{--orange:#f47b20;--black:#171717;--soft:#f4f4f4;--green:#198754;--red:#b42318}
*{box-sizing:border-box} body{font-family:Arial,Helvetica,sans-serif;margin:0;background:#f5f5f5;color:#171717}
.top{background:#171717;color:white;padding:14px 18px;border-bottom:5px solid #f47b20}
.top b{font-size:20px}.top a{color:white;margin-left:18px;text-decoration:none}
.wrap{max-width:1050px;margin:22px auto;padding:0 14px}
.card{background:white;border-radius:14px;padding:20px;box-shadow:0 2px 12px rgba(0,0,0,.08);margin-bottom:16px}
h1,h2,h3{margin-top:0}.muted{color:#666}.big{font-size:34px;font-weight:800}
.saldo{display:inline-block;background:#fff3e9;border:2px solid var(--orange);padding:9px 15px;border-radius:12px;font-weight:800}
.btn{display:inline-block;border:0;border-radius:10px;padding:12px 16px;font-size:16px;font-weight:700;cursor:pointer;text-decoration:none}
.btn-orange{background:var(--orange);color:white}.btn-black{background:#171717;color:white}.btn-green{background:var(--green);color:white}
.btn-gray{background:#e8e8e8;color:#111}.btn-red{background:var(--red);color:white}
input,select{width:100%;padding:11px;border:1px solid #bbb;border-radius:9px;font-size:16px}
label{font-weight:700;display:block;margin:10px 0 6px}
.grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(250px,1fr));gap:14px}
table{width:100%;border-collapse:collapse;background:white} th,td{padding:10px;border-bottom:1px solid #ddd;text-align:left}
th{background:#171717;color:#fff;position:sticky;top:0}.ok{color:var(--green);font-weight:800}.bad{color:var(--red);font-weight:800}
.notice{padding:12px;border-radius:9px;background:#fff3e9;border-left:5px solid var(--orange);margin:12px 0}
.center{text-align:center}.status{font-size:38px;font-weight:900}.status.ok{color:var(--green)}
.jcrc-panel{
  max-width:590px;
  margin:26px auto;
  background:#fff;
  border-radius:28px;
  overflow:hidden;
  box-shadow:0 18px 50px rgba(0,0,0,.16);
  border:1px solid #eee;
}

.jcrc-head{
  background:linear-gradient(135deg,#ff7900,#ff5c00);
  color:#fff;
  padding:30px 22px;
  text-align:center;
}

.jcrc-mark{
  width:92px;
  height:92px;
  border-radius:50%;
  background:#fff;
  color:#f47b20;
  display:flex;
  align-items:center;
  justify-content:center;
  margin:0 auto 16px;
  font-size:24px;
  font-weight:900;
  border:6px solid rgba(255,255,255,.45);
  box-shadow:0 7px 20px rgba(0,0,0,.20);
}

.jcrc-head h1{
  font-size:32px;
  margin-bottom:7px !important;
}

.jcrc-body{
  padding:28px;
  background:linear-gradient(#fff,#fffaf6);
}

.request-number{
  font-size:70px;
  font-weight:900;
  color:#f47b20;
  line-height:1;
  margin:12px 0 5px;
}

.info-box{
  background:#fff;
  border-radius:18px;
  padding:18px;
  margin:20px 0;
  text-align:left;
  border:1px solid #eee;
  box-shadow:0 4px 15px rgba(0,0,0,.06);
  font-size:17px;
  line-height:1.7;
}

.action-btn{
  width:100%;
  font-size:20px;
  padding:17px;
  border-radius:16px;
  font-weight:900;
  box-shadow:0 5px 14px rgba(0,0,0,.12);
}

.btn-green.action-btn{
  background:linear-gradient(135deg,#ff7900,#ff5c00);
  color:white;
}

.btn-red.action-btn{
  background:white;
  color:#c92d2d;
  border:2px solid #e34b4b;
}

.share-card{
  max-width:560px;
  margin:30px auto;
  text-align:center;
}

.share-card img{
  width:min(300px,82vw);
  height:auto;
  border:10px solid #fff;
  box-shadow:0 5px 20px rgba(0,0,0,.12);
  border-radius:14px;
}
.share-card{max-width:560px;margin:30px auto;text-align:center}.share-card img{width:min(300px,82vw);height:auto;border:10px solid #fff;box-shadow:0 5px 20px rgba(0,0,0,.12);border-radius:14px}
@media(max-width:700px){table{font-size:13px}.hide-mobile{display:none}.big{font-size:28px}}
"""


def page(title, body, script="", head_extra=""):
    nav = ""
    if is_admin():
        nav = """<a href="/admin">Inicio</a><a href="/admin/qrs">QR para imprimir</a><a href="/admin/history">Historial</a><a href="/logout">Salir</a>"""
    return render_template_string(
        """<!doctype html><html lang="es"><head><meta charset="utf-8">
        <meta name="viewport" content="width=device-width,initial-scale=1">
        <meta name="robots" content="noindex,nofollow">
        {{head_extra|safe}}
        <title>{{title}}</title><style>{{css}}</style></head>
        <body><div class="top"><b>JCRC · Vouchers Pileta</b>{{nav|safe}}</div>
        <div class="wrap">{{body|safe}}</div>{{script|safe}}</body></html>""",
        title=title, css=BASE_CSS, nav=nav, body=body, script=script, head_extra=head_extra
    )
@app.get("/jcrc_logo.png")
def jcrc_logo():
    return send_file("jcrc_logo.png", mimetype="image/png")

@app.get("/")
def home():
    return redirect(url_for("admin_home") if is_admin() else url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    error = ""
    if request.method == "POST":
        pin = request.form.get("pin", "")
        if ADMIN_PIN and secrets.compare_digest(pin, ADMIN_PIN):
            session.clear()
            session["admin"] = True
            session.permanent = True
            nxt = request.form.get("next") or "/admin"
            return redirect(nxt if nxt.startswith("/") or nxt.startswith(public_base()) else "/admin")
        error = '<div class="notice bad">PIN incorrecto.</div>'
    nxt = session.get("next", "/admin")
    body = f"""
    <div class="card" style="max-width:430px;margin:50px auto">
      <h1>Ingreso recepción</h1>
      <p class="muted">Sistema de vouchers · Temporada 2026/27</p>
      {error}
      <form method="post">
        <input type="hidden" name="next" value="{nxt}">
        <label>PIN de recepción</label><input type="password" name="pin" inputmode="numeric" autofocus required>
        <br><br><button class="btn btn-orange" style="width:100%">Ingresar</button>
      </form>
    </div>"""
    return page("Ingreso", body)


@app.get("/logout")
def logout():
    session.clear()
    return redirect("/login")


@app.get("/admin")
@admin_required
def admin_home():
    try:
        state, _ = load_state()
        items = []
        q = request.args.get("q", "").strip().lower()
        for mid, member in state.get("members", {}).items():
            if not member.get("activo", True):
                continue
            p = decrypt_member(member)
            hay = f"{p.get('nombre','')} {p.get('socio','')}".lower()
            if q and q not in hay:
                continue
            items.append((p.get("nombre",""), mid, member, p))
        items.sort(key=lambda x: x[0].lower())
        rows = []
        for _, mid, m, p in items:
            phone = p.get("telefono_wa") or ""
            phone_html = '<span class="ok">OK</span>' if phone else '<span class="bad">FALTA</span>'
            rows.append(f"""
              <tr>
                <td><b>{p.get('nombre','')}</b><br><span class="muted">Socio {p.get('socio','')}</span></td>
                <td>{p.get('categoria','')}</td>
                <td><b>{m.get('saldo',0)}</b></td>
                <td>{phone_html}</td>
                <td><a class="btn btn-orange" href="/v/{mid}">Abrir</a>
                    <a class="btn btn-gray" href="/qr/{mid}.png?download=1">QR</a>
                    {f'<a class="btn btn-green" href="/admin/send-voucher/{mid}" target="_blank" rel="noopener">Enviar</a>' if phone else '<span class="muted">Sin tel.</span>'}</td>
              </tr>""")
        body = f"""
        <div class="card">
          <h1>Recepción · Vouchers</h1>
          <form method="get"><div style="display:flex;gap:8px">
            <input name="q" value="{request.args.get('q','')}" placeholder="Buscar por titular o Nº de socio">
            <button class="btn btn-black">Buscar</button>
          </div></form>
        </div>
        <div class="card">
          <h3>Actualizar desde Excel</h3>
          <p class="muted">Elegí tu archivo <b>temporada 2026-27.xlsx</b>. Se actualizan todas las filas y campos A:H. La columna Invitaciones funciona como carga pendiente: ponerla en 0 no borra el saldo; una nueva cantidad después de 0 se suma una sola vez al mismo voucher.</p>
          <form method="post" action="/admin/upload-excel" enctype="multipart/form-data">
            <div style="display:flex;gap:10px;align-items:end;flex-wrap:wrap">
              <div style="flex:1;min-width:260px"><label>Archivo Excel</label><input type="file" name="excel" accept=".xlsx,.xlsm" required></div>
              <button class="btn btn-orange">ACTUALIZAR EXCEL</button>
            </div>
          </form>
        </div>
        <div class="card" style="overflow:auto">
          <table><thead><tr><th>Titular</th><th>Categoría</th><th>Saldo</th><th>Tel.</th><th>Acción</th></tr></thead>
          <tbody>{''.join(rows)}</tbody></table>
        </div>"""
        if not GITHUB_TOKEN:
            body = '<div class="notice bad">Falta GITHUB_DATA_TOKEN: los cambios no serán persistentes.</div>' + body
        return page("Recepción", body)
    except Exception as e:
        return page("Error", f'<div class="card"><h2>Error</h2><p class="bad">{e}</p></div>'), 500



@app.post("/admin/upload-excel")
@admin_required
def upload_excel():
    f = request.files.get("excel")
    if not f or not f.filename:
        return page("Actualizar Excel", '<div class="card"><h2>Falta el archivo</h2><a class="btn btn-gray" href="/admin">Volver</a></div>'), 400
    if not f.filename.lower().endswith((".xlsx", ".xlsm")):
        return page("Actualizar Excel", '<div class="card"><h2>Formato no válido</h2><p>Elegí un archivo .xlsx.</p><a class="btn btn-gray" href="/admin">Volver</a></div>'), 400
    try:
        excel_rows = parse_excel_members(f)

        def mutate(state):
            return sync_excel_into_state(state, excel_rows)

        result = update_state(mutate)
        body = f"""
        <div class="card center" style="max-width:650px;margin:30px auto">
          <div class="status ok">EXCEL ACTUALIZADO</div>
          <p><b>{result['rows']}</b> titulares actualizados desde A:H.</p>
          <p>Nuevos: <b>{result['added']}</b> · Actualizados: <b>{result['updated']}</b> · Ya no presentes: <b>{result['deactivated']}</b></p>
          <p>Nuevas invitaciones acreditadas: <b>{result['credited_qty']}</b> en <b>{result['credited_members']}</b> titular(es).</p>
          <div class="notice">La columna Invitaciones no reemplaza el saldo: al ponerla en 0 el voucher conserva lo que tenga. Cuando después cargás una nueva cantidad, se suma una sola vez.</div>
          <a class="btn btn-orange" href="/admin">Volver a recepción</a>
        </div>"""
        return page("Excel actualizado", body)
    except Exception as e:
        return page("Error al actualizar", f'<div class="card"><h2>No pude actualizar el Excel</h2><p class="bad">{e}</p><a class="btn btn-gray" href="/admin">Volver</a></div>'), 400


@app.get("/admin/send-voucher/<mid>")
@admin_required
def send_voucher(mid):
    state, _ = load_state()
    m = state.get("members", {}).get(mid)
    if not m or not m.get("activo", True):
        abort(404)
    p = decrypt_member(m)
    phone = p.get("telefono_wa") or ""
    if not phone:
        return page(
            "Falta teléfono",
            '<div class="card"><h2>Falta teléfono válido</h2><p>Actualizá el teléfono del titular y volvé a intentar.</p><a class="btn btn-gray" href="/admin">Volver</a></div>',
        ), 400

    share_url = voucher_share_url(mid)
    saldo = int(m.get("saldo", 0))
    msg = (
        f"Hola {p.get('nombre','')}. 👋\n\n"
        f"Jockey Club Río Cuarto te envía tu voucher de invitaciones "
        f"para la temporada 2026/27.\n\n"
        f"Saldo disponible: {saldo} invitación{'es' if saldo != 1 else ''}.\n\n"
        f"Tu voucher / QR:\n{share_url}\n\n"
        f"Podés conservar este enlace y compartir el QR con tu grupo."
    )
    return redirect(f"https://wa.me/{phone}?text={quote(msg)}")


@app.get("/voucher/<mid>/<sig>")
def public_voucher(mid, sig):
    if not valid_voucher_signature(mid, sig):
        abort(404)
    state, _ = load_state()
    m = state.get("members", {}).get(mid)
    if not m or not m.get("activo", True):
        abort(404)
    p = decrypt_member(m)
    saldo = int(m.get("saldo", 0))
    body = f"""
    <div class="jcrc-panel">
      <div class="jcrc-head">
       <img src="/jcrc_logo.png" alt="Jockey Club Río Cuarto"
style="width:105px;height:105px;object-fit:contain;border-radius:50%;
background:white;padding:5px;margin:0 auto 16px;display:block;
box-shadow:0 7px 20px rgba(0,0,0,.20)">
        <h1 style="margin-bottom:4px">Voucher de invitaciones</h1>
        <div>Temporada 2026/27</div>
      </div>
      <div class="jcrc-body center">
        <div class="big">{p.get('nombre','')}</div>
        <p class="muted">Socio Nº {p.get('socio','')}</p>
        <img src="/voucher/{mid}/{sig}/qr.png" alt="QR del voucher"
             style="width:min(300px,82vw);height:auto;margin:8px auto 18px;display:block">
        <div class="saldo">Saldo disponible: {saldo}</div>
        <p class="muted" style="margin-top:18px">
          Presentá este QR en recepción. El ingreso queda sujeto a la autorización del titular.
        </p>
      </div>
    </div>"""
    return page("Voucher JCRC", body)


@app.get("/voucher/<mid>/<sig>/qr.png")
def public_voucher_qr(mid, sig):
    from PIL import Image

    if not valid_voucher_signature(mid, sig):
        abort(404)

    state, _ = load_state()
    if mid not in state.get("members", {}):
        abort(404)

    url = f"{public_base()}/v/{mid}"

    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=9,
        border=3
    )

    qr.add_data(url)
    qr.make(fit=True)

    img = qr.make_image(
        fill_color="black",
        back_color="white"
    ).convert("RGBA")

    logo = Image.open("jcrc_logo.png").convert("RGBA")

    qr_w, qr_h = img.size
    logo_size = int(qr_w * 0.18)
    logo.thumbnail((logo_size, logo_size))

    pos = (
        (qr_w - logo.width) // 2,
        (qr_h - logo.height) // 2
    )

    img.alpha_composite(logo, pos)

    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)

    return send_file(
        buf,
        mimetype="image/png",
        download_name=f"voucher-{mid}.png"
    )
@app.route("/v/<mid>", methods=["GET"])
@admin_required
def voucher(mid):
    state, _ = load_state()
    m = state.get("members", {}).get(mid)
    if not m:
        abort(404)
    p = decrypt_member(m)
    saldo = int(m.get("saldo", 0))
    phone = p.get("telefono_wa") or ""
    phone_note = f'<span class="ok">WhatsApp: +{phone}</span>' if phone else '<span class="bad">Falta teléfono válido</span>'
    qty_options = "".join(f'<option value="{i}">{i}</option>' for i in range(1, max(saldo, 0)+1))
    form = ""
    if saldo > 0 and phone:
        form = f"""
        <form method="post" action="/v/{mid}/request">
          <label>Personas que quieren ingresar</label>
          <select name="qty">{qty_options}</select><br><br>
          <button class="btn btn-orange" style="width:100%">Crear solicitud de autorización</button>
        </form>"""
    elif saldo <= 0:
        form = '<div class="notice bad">Sin invitaciones disponibles.</div>'
    else:
        form = '<div class="notice">Cargá un teléfono para poder abrir WhatsApp.</div>'
    body = f"""
    <div class="card">
      <a href="/admin">← Volver</a><br><br>
      <div class="grid">
        <div><div class="muted">Titular</div><div class="big">{p.get('nombre','')}</div>
          <p><b>Nº socio:</b> {p.get('socio','')}<br><b>Categoría:</b> {p.get('categoria','')}<br>{phone_note}</p></div>
        <div class="center"><div class="muted">Saldo disponible</div><div class="big">{saldo}</div>
          <span class="saldo">Total cargadas: {m.get('invitaciones_iniciales',0)}</span></div>
      </div>
      <hr style="border:0;border-top:1px solid #ddd;margin:20px 0">
      {form}
    </div>
    <div class="card">
      <h3>Editar teléfono / saldo</h3>
      <form method="post" action="/admin/member/{mid}/edit">
        <div class="grid">
          <div><label>Teléfono</label><input name="phone" value="{p.get('telefono_original','')}" placeholder="Ej. 358 4012345"></div>
          <div><label>Saldo actual</label><input name="saldo" type="number" min="0" value="{saldo}"></div>
        </div><br><button class="btn btn-gray">Guardar corrección</button>
      </form>
    </div>"""
    return page(p.get("nombre","Voucher"), body)


@app.post("/admin/member/<mid>/edit")
@admin_required
def edit_member(mid):
    phone_input = request.form.get("phone", "").strip()
    saldo_input = request.form.get("saldo", "").strip()

    def mutate(state):
        m = state["members"].get(mid)
        if not m:
            raise ValueError("Titular inexistente")
        p = decrypt_member(m)
        p["telefono_original"] = phone_input
        p["telefono_wa"] = normalize_phone(phone_input)
        m["payload"] = encrypt_payload(p)
        if saldo_input != "":
            m["saldo"] = max(0, int(saldo_input))
        state.setdefault("history", []).append({
            "at": now_iso(), "type": "admin_edit", "member_id": mid,
            "saldo": m["saldo"]
        })
        state["history"] = state["history"][-500:]
    update_state(mutate)
    return redirect(f"/v/{mid}")


@app.post("/v/<mid>/request")
@admin_required
def create_request(mid):
    qty = int(request.form.get("qty", "0"))
    token = secrets.token_urlsafe(24)

    def mutate(state):
        m = state.get("members", {}).get(mid)
        if not m:
            raise ValueError("Titular inexistente")
        saldo = int(m.get("saldo", 0))
        if qty < 1 or qty > saldo:
            raise ValueError("Cantidad inválida o saldo insuficiente")
        p = decrypt_member(m)
        if not p.get("telefono_wa"):
            raise ValueError("Falta teléfono válido")
        state.setdefault("requests", {})[token] = {
            "member_id": mid,
            "qty": qty,
            "status": "pending",
            "created_at": now_iso(),
            "created_ts": now_ts(),
        }
        return p, saldo

    p, saldo = update_state(mutate)
    auth_url = f"{public_base()}/a/{token}"
    msg = (
    f"🟠 *JOCKEY CLUB RÍO CUARTO*\n"
    f"*Solicitud de autorización de ingreso*\n\n"
    f"Titular: {p.get('nombre','')}\n"
    f"Nº socio: {p.get('socio','')}\n"
    f"Ingreso solicitado: {qty} persona{'s' if qty != 1 else ''}\n"
    f"Saldo disponible: {saldo}\n\n"
    f"👇 *ABRÍ LA TARJETA DE AUTORIZACIÓN*\n"
    f"{auth_url}\n\n"
    f"Ahí vas a ver dos botones grandes:\n"
    f"✅ *Aceptar*   ❌ *Rechazar*\n\n"
    f"⏱️ La solicitud vence en 30 minutos."
    )
    wa_url = f"https://wa.me/{p['telefono_wa']}?text={quote(msg)}"
    body = f"""
    <div class="card center">
      <h1>Solicitud creada</h1>
      <p><b>{p.get('nombre','')}</b> · {qty} persona{'s' if qty != 1 else ''}</p>
      <a class="btn btn-green" href="{wa_url}" target="_blank" rel="noopener">ENVIAR SOLICITUD POR WHATSAPP</a>
      <p class="muted">Después de enviar, dejá esta pantalla abierta. Se actualizará sola.</p>
      <div id="st" class="status">ESPERANDO...</div>
      <p id="detail">Saldo actual: {saldo}</p>
      <a class="btn btn-gray" href="/v/{mid}">Volver al voucher</a>
    </div>"""
    script = f"""<script>
      const poll=()=>fetch('/api/request/{token}').then(r=>r.json()).then(x=>{{
        const el=document.getElementById('st'), d=document.getElementById('detail');
        if(x.status==='approved'){{el.textContent='AUTORIZADO';el.className='status ok';d.textContent='Nuevo saldo: '+x.saldo;}}
        else if(x.status==='rejected'){{el.textContent='RECHAZADO';el.className='status bad';}}
        else if(x.status==='expired'){{el.textContent='VENCIDO';el.className='status bad';}}
        else setTimeout(poll,2500);
      }}).catch(()=>setTimeout(poll,4000)); poll();
    </script>"""
    return page("Solicitud", body, script)


@app.get("/api/request/<token>")
@admin_required
def request_status(token):
    state, _ = load_state()
    req = state.get("requests", {}).get(token)
    if not req:
        return jsonify({"status": "missing"}), 404
    status = req.get("status", "pending")
    if status == "pending" and now_ts() - int(req.get("created_ts", 0)) > 1800:
        status = "expired"
    m = state.get("members", {}).get(req["member_id"], {})
    return jsonify({"status": status, "saldo": m.get("saldo", 0)})


@app.route("/a/<token>", methods=["GET"])
def @app.route("/a/<token>", methods=["GET"])
def authorize_page(token):
    state, _ = load_state()
    req = state.get("requests", {}).get(token)

    if not req:
        return page(
            "Solicitud inexistente",
            '<div class="card center"><h2>Solicitud inexistente</h2></div>'
        ), 404

    m = state.get("members", {}).get(req["member_id"])
    if not m:
        abort(404)

    p = decrypt_member(m)
    status = req.get("status", "pending")

    if status == "pending" and now_ts() - int(req.get("created_ts", 0)) > 1800:
        status = "expired"

    logo = """
    <img src="/jcrc_logo.png"
         alt="Jockey Club Río Cuarto"
         style="width:105px;height:105px;object-fit:contain;border-radius:50%;
         background:white;padding:5px;margin:0 auto 16px;display:block;
         box-shadow:0 7px 20px rgba(0,0,0,.20)">
    """

    if status == "approved":
        body = f"""
        <div class="jcrc-panel">
          <div class="jcrc-head">
            {logo}
            <h1>Autorización</h1>
          </div>
          <div class="jcrc-body center">
            <div class="status ok">✓ AUTORIZADO</div>
            <p style="font-size:18px">{req["qty"]} persona(s).</p>
            <p>Recepción ya recibió tu respuesta.</p>
          </div>
        </div>
        """

    elif status == "rejected":
        body = f"""
        <div class="jcrc-panel">
          <div class="jcrc-head">
            {logo}
            <h1>Autorización</h1>
          </div>
          <div class="jcrc-body center">
            <div class="status bad">✕ RECHAZADO</div>
            <p>La solicitud fue rechazada.</p>
            <p>Recepción ya recibió tu respuesta.</p>
          </div>
        </div>
        """

    elif status == "expired":
        body = f"""
        <div class="jcrc-panel">
          <div class="jcrc-head">
            {logo}
            <h1>Autorización</h1>
          </div>
          <div class="jcrc-body center">
            <div class="status bad">VENCIDO</div>
            <p>Pedí a recepción una nueva solicitud.</p>
          </div>
        </div>
        """

    else:
        body = f"""
        <div class="jcrc-panel">
          <div class="jcrc-head">
            {logo}
            <h1 style="margin-bottom:5px">Autorización de ingreso</h1>
            <div>Jockey Club Río Cuarto</div>
          </div>

          <div class="jcrc-body center">

            <div style="display:inline-block;background:white;border:1px solid #ddd;
            padding:9px 15px;border-radius:20px;
            box-shadow:0 3px 10px rgba(0,0,0,.08);
            font-weight:700;margin-bottom:18px">
              🟢 Válido desde WhatsApp
            </div>

            <p style="font-size:17px">
              Recepción solicita autorización para el ingreso de:
            </p>

            <div class="request-number">{req['qty']}</div>

            <div style="font-size:19px;font-weight:900;margin-bottom:12px">
              persona{'s' if req['qty'] != 1 else ''}
            </div>

            <div style="color:#f47b20;font-weight:900;margin-bottom:20px">
              ◷ Válido por 30 minutos
            </div>

            <div class="info-box">
              <span style="color:#777">Titular</span><br>
              <b style="font-size:19px">{p.get('nombre','')}</b>

              <hr style="border:0;border-top:1px solid #eee">

              <span style="color:#777">Nº de socio</span><br>
              <b style="font-size:19px">{p.get('socio','')}</b>

              <hr style="border:0;border-top:1px solid #eee">

              <span style="color:#777">Saldo disponible</span><br>
              <b style="font-size:19px">
                {m.get('saldo',0)} invitación{'es' if int(m.get('saldo',0)) != 1 else ''}
              </b>
            </div>

            <p style="font-size:19px;font-weight:900">
              ¿Autorizás este ingreso?
            </p>

            <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px">
              <form method="post" action="/a/{token}/approve">
                <button class="btn btn-green action-btn">✓ Aceptar</button>
              </form>

              <form method="post" action="/a/{token}/reject">
                <button class="btn btn-red action-btn">✕ Rechazar</button>
              </form>
            </div>

          </div>
        </div>
        """

    og_title = "Jockey Club Río Cuarto · Autorización de ingreso"
    og_desc = (
        f"Solicitud para autorizar el ingreso de {req['qty']} "
        f"persona{'s' if req['qty'] != 1 else ''}. Tocá para Aceptar o Rechazar."
    )

    head_extra = f"""
    <meta property="og:type" content="website">
    <meta property="og:title" content="{og_title}">
    <meta property="og:description" content="{og_desc}">
    <meta property="og:url" content="{request.url}">
    """

    return page("Autorizar ingreso", body, head_extra=head_extra)
@app.post("/a/<token>/approve")
def approve(token):
    def mutate(state):
        req = state.get("requests", {}).get(token)
        if not req:
            return "missing", None
        if req.get("status") != "pending":
            return req.get("status"), state["members"][req["member_id"]].get("saldo", 0)
        if now_ts() - int(req.get("created_ts", 0)) > 1800:
            req["status"] = "expired"
            return "expired", state["members"][req["member_id"]].get("saldo", 0)
        m = state["members"][req["member_id"]]
        qty = int(req["qty"])
        if int(m.get("saldo", 0)) < qty:
            req["status"] = "insufficient"
            return "insufficient", m.get("saldo", 0)
        m["saldo"] = int(m.get("saldo", 0)) - qty
        req["status"] = "approved"
        req["resolved_at"] = now_iso()
        state.setdefault("history", []).append({
            "at": now_iso(), "type": "approved", "member_id": req["member_id"],
            "qty": qty, "saldo": m["saldo"]
        })
        state["history"] = state["history"][-500:]
        return "approved", m["saldo"]
    status, saldo = update_state(mutate)
    return redirect(f"/a/{token}")


@app.post("/a/<token>/reject")
def reject(token):
    def mutate(state):
        req = state.get("requests", {}).get(token)
        if not req:
            return
        if req.get("status") == "pending":
            req["status"] = "rejected"
            req["resolved_at"] = now_iso()
            state.setdefault("history", []).append({
                "at": now_iso(), "type": "rejected", "member_id": req["member_id"],
                "qty": req["qty"]
            })
            state["history"] = state["history"][-500:]
    update_state(mutate)
    return redirect(f"/a/{token}")


@app.get("/qr/<mid>.png")
@admin_required
def qr_png(mid):
    from PIL import Image

    state, _ = load_state()
    if mid not in state.get("members", {}):
        abort(404)

    url = f"{public_base()}/v/{mid}"

    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=9,
        border=3
    )

    qr.add_data(url)
    qr.make(fit=True)

    img = qr.make_image(
        fill_color="black",
        back_color="white"
    ).convert("RGBA")

    logo = Image.open("jcrc_logo.png").convert("RGBA")

    qr_w, qr_h = img.size
    logo_size = int(qr_w * 0.18)
    logo.thumbnail((logo_size, logo_size))

    pos = (
        (qr_w - logo.width) // 2,
        (qr_h - logo.height) // 2
    )

    img.alpha_composite(logo, pos)

    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)

    return send_file(
        buf,
        mimetype="image/png",
        as_attachment=request.args.get("download") == "1",
        download_name=f"voucher-{mid}.png"
    )
@app.get("/admin/qrs")
@admin_required
def print_qrs():
    state, _ = load_state()
    cards = []
    for mid, m in state.get("members", {}).items():
        if not m.get("activo", True):
            continue
        p = decrypt_member(m)
        cards.append((p.get("nombre",""), f"""
          <div class="qr-card">
            <div class="brand">JCRC · TEMPORADA 2026/27</div>
            <img src="/qr/{mid}.png">
            <div class="name">{p.get('nombre','')}</div>
            <div>Socio {p.get('socio','')}</div>
            <div class="small">Voucher de invitaciones</div>
          </div>"""))
    cards.sort(key=lambda x: x[0].lower())
    body = '<div class="card no-print"><h1>QR para imprimir</h1><button class="btn btn-orange" onclick="window.print()">Imprimir</button></div><div class="qr-grid">' + "".join(x[1] for x in cards) + '</div>'
    extra = """<style>
    .qr-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:12px}.qr-card{background:#fff;border:2px solid #171717;border-top:8px solid #f47b20;border-radius:10px;text-align:center;padding:12px;break-inside:avoid}.qr-card img{width:170px;height:170px}.brand{font-weight:900}.name{font-size:18px;font-weight:800;margin-top:6px}.small{font-size:12px;color:#555}
    @media print{.top,.no-print{display:none!important}.wrap{max-width:none;margin:0;padding:0}.qr-grid{grid-template-columns:repeat(3,1fr);gap:8px}.qr-card{box-shadow:none}}
    </style>"""
    return page("QR para imprimir", body, extra)


@app.get("/admin/history")
@admin_required
def history():
    state, _ = load_state()
    rows = []
    for h in reversed(state.get("history", [])[-200:]):
        m = state.get("members", {}).get(h.get("member_id"), {})
        try:
            p = decrypt_member(m) if m else {}
        except Exception:
            p = {}
        rows.append(f"<tr><td>{h.get('at','')[:19].replace('T',' ')}</td><td>{h.get('type','')}</td><td>{p.get('nombre','')}</td><td>{h.get('qty','')}</td><td>{h.get('saldo','')}</td></tr>")
    body = f'<div class="card"><h1>Historial</h1><table><tr><th>Fecha</th><th>Evento</th><th>Titular</th><th>Cant.</th><th>Saldo</th></tr>{"".join(rows)}</table></div>'
    return page("Historial", body)


@app.get("/health")
def health():
    try:
        state, _ = load_state()
        return jsonify({"status": "ok", "members": len(state.get("members", {})), "github_persistence": bool(GITHUB_TOKEN)})
    except Exception as e:
        return jsonify({"status": "error", "error": str(e)}), 500


@app.get("/privacy")
def privacy():
    return """<!doctype html><html lang="es"><head><meta charset="utf-8"><title>Política de Privacidad - Jockey Club Río Cuarto</title></head>
    <body style="font-family:Arial;max-width:800px;margin:40px auto;line-height:1.5"><h1>Política de Privacidad</h1><h2>Jockey Club Río Cuarto</h2>
    <p>Esta aplicación gestiona vouchers de temporada, invitaciones y autorizaciones de ingreso.</p>
    <p>Los datos se utilizan únicamente para administrar el servicio, validar accesos y gestionar autorizaciones.</p>
    <p>No se venden ni alquilan datos personales a terceros.</p>
    <p>Los usuarios pueden solicitar acceso, modificación o eliminación de sus datos mediante los canales oficiales del Jockey Club Río Cuarto.</p></body></html>"""


# Se conservan los endpoints del webhook ya configurado en Meta, aunque esta versión no depende de la API.
@app.get("/webhook")
def verify_webhook():
    mode = request.args.get("hub.mode")
    token = request.args.get("hub.verify_token")
    challenge = request.args.get("hub.challenge")
    if mode == "subscribe" and token == VERIFY_TOKEN:
        return challenge or "", 200
    return "Verification failed", 403


@app.post("/webhook")
def receive_webhook():
    payload = request.get_json(silent=True) or {}
    print("WHATSAPP WEBHOOK:", payload, flush=True)
    return jsonify({"status": "ok"}), 200


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port)
